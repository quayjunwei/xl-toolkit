"""Generate cropped satellite thumbnails for features and embed them into an Excel workbook.

For each data row, the script matches the row's ID against the `osm_id` field of a
QGIS layer, takes that feature's geometry centroid (falling back to a manual lat/lng
column if no match is found in the layer), downloads and stitches nearby satellite
tiles around that point, crops the result to a square centered on the point with a
red crosshair marker, then inserts the thumbnail image back into the workbook next
to the row it came from.

Requirements for the input Excel file (`excel_path`):
    - Row 1 must contain column headers; data begins on row 2.
    - A column headed `id_column_header` (default "osm_id") whose values match the
      `osm_id` field of features in the QGIS layer (`layer_name`).
    - Optionally, a column headed `latlng_column_header` (default "lat_lng") holding
      "latitude,longitude" strings, used only for rows whose ID has no match in the
      QGIS layer.
    - The column at `thumbnail_column_letter` (default "H") should be left empty —
      that's where each row's thumbnail image is inserted.
"""

import os
import math
import json
import requests
import time
from io import BytesIO
from PIL import Image, ImageDraw
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from concurrent.futures import ThreadPoolExecutor, as_completed
from qgis.core import QgsProject, QgsCoordinateReferenceSystem, QgsCoordinateTransform

# Configuration
layer_name = "<QGIS_LAYER_NAME>"
zoom = 18
tile_size = 256
crop_size = 256
grid_size = 3  # Multi-tile stitching for better context
stitched_size = tile_size * grid_size
tile_url_template = "<TILE_SERVER_URL_TEMPLATE>"  # must contain {x}, {y}, {z} placeholders, e.g. an XYZ satellite tile server
output_base = "<PATH_TO_OUTPUT_IMAGE_FOLDER>"
excel_path = "<PATH_TO_INPUT_EXCEL_FILE>"
output_excel = "<PATH_TO_OUTPUT_EXCEL_FILE>"
thumbnail_column_letter = "H"
id_column_header = "osm_id"  # column matched by header name, not index
latlng_column_header = "lat_lng"  # fallback column, only used if osm_id has no match in the layer
max_threads_tiles = 3  # Reduced for stability
cache_file = os.path.join(output_base, "tile_cache.json")

os.makedirs(output_base, exist_ok=True)

# Load QGIS layer
layer = QgsProject.instance().mapLayersByName(layer_name)[0]
crs_src = layer.crs()
crs_dest = QgsCoordinateReferenceSystem("EPSG:4326")
xform = QgsCoordinateTransform(crs_src, crs_dest, QgsProject.instance())

# Load Excel
wb = load_workbook(excel_path)
ws = wb.active


def find_column_by_header(ws, header_name):
    """Find a column by its row-1 header text.

    Args:
        ws: The worksheet to search.
        header_name: Header text to match, case-insensitively.

    Returns:
        The 1-indexed column number, or None if no header matches.
    """
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        if cell_value and str(cell_value).strip().lower() == header_name.lower():
            return col_idx
    return None


id_column_index = find_column_by_header(ws, id_column_header)
if id_column_index is None:
    raise ValueError(f"Could not find a column header named '{id_column_header}' in {excel_path}")

latlng_column_index = find_column_by_header(ws, latlng_column_header)
if latlng_column_index is None:
    print(f"No '{latlng_column_header}' column found — rows without a matching {id_column_header} in the layer will be skipped instead of falling back to manual lat/lng")

feature_dict = {str(f["osm_id"]): f for f in layer.getFeatures()}

headers = {"User-Agent": "Mozilla/5.0"}
tile_cache = {}
if os.path.exists(cache_file):
    with open(cache_file, "r", encoding="utf-8") as f:
        tile_cache = json.load(f)


def save_cache():
    """Persist the in-memory tile cache to cache_file as JSON."""
    with open(cache_file, "w", encoding="utf-8") as f:
        json.dump(tile_cache, f, ensure_ascii=False, indent=2)


def latlon_to_tile_coords(lat, lon, zoom):
    """Convert a lat/lon in degrees to fractional slippy-map tile coordinates.

    Args:
        lat: Latitude in degrees.
        lon: Longitude in degrees.
        zoom: Slippy-map zoom level.

    Returns:
        A tuple of (x, y) fractional tile coordinates.
    """
    lat_rad = math.radians(lat)
    n = 2.0**zoom
    x = (lon + 180.0) / 360.0 * n
    y = (
        (1.0 - math.log(math.tan(lat_rad) + 1.0 / math.cos(lat_rad)) / math.pi)
        / 2.0
        * n
    )
    return x, y


def get_tile(tile_x, tile_y, zoom):
    """Fetch a single map tile, using the on-disk cache if already downloaded.

    Args:
        tile_x: Tile column index.
        tile_y: Tile row index.
        zoom: Slippy-map zoom level.

    Returns:
        A PIL Image of the tile, or None if the download failed after retries.
    """
    key = f"{tile_x}_{tile_y}_{zoom}"
    if key in tile_cache:
        return Image.open(tile_cache[key])

    tile_url = tile_url_template.format(x=tile_x, y=tile_y, z=zoom)
    for attempt in range(3):  # Retry logic
        try:
            response = requests.get(tile_url, headers=headers, timeout=10)
            if response.status_code == 200:
                img = Image.open(BytesIO(response.content)).convert("RGB")
                path = os.path.join(output_base, f"tile_{key}.jpg")
                img.save(path)
                tile_cache[key] = path
                return img
        except requests.exceptions.SSLError:
            print(f"SSL error on attempt {attempt+1} for tile {key}, retrying...")
            time.sleep(2)
    print(f"Failed to download tile {key}")
    return None


def process_row(row_idx, total_rows):
    """Build and save a cropped, cross-marked satellite thumbnail for one spreadsheet row.

    Args:
        row_idx: 1-indexed worksheet row to process.
        total_rows: Total row count, used only for progress messages.

    Returns:
        A tuple of (row_idx, output_path) on success, or None if the row was skipped.
    """
    place_id = str(ws.cell(row=row_idx, column=id_column_index).value)
    output_path = os.path.join(output_base, f"{place_id}.jpeg")
    if os.path.exists(output_path):
        print(f"Skipping row {row_idx}/{total_rows}: image already exists")
        return None

    if place_id in feature_dict:
        geom = feature_dict[place_id].geometry().centroid().asPoint()
        wgs_point = xform.transform(geom)
        lat, lon = wgs_point.y(), wgs_point.x()
    else:
        if latlng_column_index is None:
            print(f"Skipping row {row_idx}/{total_rows}: {id_column_header} not in layer and no lat_lng fallback column")
            return
        latlng = ws.cell(row=row_idx, column=latlng_column_index).value
        if not latlng or not isinstance(latlng, str) or "," not in latlng:
            print(f"Skipping row {row_idx}/{total_rows}")
            return
        try:
            lat, lon = map(float, latlng.split(","))
        except:
            print(f"Invalid lat_lng format at row {row_idx}")
            return

    print(f"Processing row {row_idx}/{total_rows}")

    x_tile_float, y_tile_float = latlon_to_tile_coords(lat, lon, zoom)
    half_grid = grid_size // 2
    top_left_x = int(x_tile_float) - half_grid
    top_left_y = int(y_tile_float) - half_grid

    stitched_img = Image.new("RGB", (stitched_size, stitched_size))

    with ThreadPoolExecutor(max_workers=max_threads_tiles) as pool:
        futures = {
            pool.submit(get_tile, x, y, zoom): (dx, dy)
            for dx, x in enumerate(range(top_left_x, top_left_x + grid_size))
            for dy, y in enumerate(range(top_left_y, top_left_y + grid_size))
        }
        for future in as_completed(futures):
            img = future.result()
            dx, dy = futures[future]
            if img:
                stitched_img.paste(img, (dx * tile_size, dy * tile_size))

    pixel_x = int((x_tile_float - top_left_x) * tile_size)
    pixel_y = int((y_tile_float - top_left_y) * tile_size)
    left = pixel_x - crop_size // 2
    upper = pixel_y - crop_size // 2
    right = pixel_x + crop_size // 2
    lower = pixel_y + crop_size // 2
    cropped_img = stitched_img.crop((left, upper, right, lower))

    draw = ImageDraw.Draw(cropped_img)

    cross_size = 8
    line_width = 2
    center_x = crop_size // 2
    center_y = crop_size // 2

    draw.line(
        (center_x - cross_size, center_y, center_x + cross_size, center_y),
        fill="red",
        width=line_width,
    )
    draw.line(
        (center_x, center_y - cross_size, center_x, center_y + cross_size),
        fill="red",
        width=line_width,
    )

    cropped_img.save(output_path)
    return row_idx, output_path


# Process rows
results = []
total_rows = ws.max_row
for row_idx in range(2, total_rows + 1):
    res = process_row(row_idx, total_rows)
    if res:
        results.append(res)

# Insert images into Excel (first occurrence of each osm_id only)
seen_ids = set()

for row_idx in range(2, total_rows + 1):
    place_id = str(ws.cell(row=row_idx, column=id_column_index).value)

    if place_id in seen_ids:
        continue  # Skip duplicate osm_ids
    seen_ids.add(place_id)

    img_path = os.path.join(output_base, f"{place_id}.jpeg")
    if os.path.exists(img_path):
        try:
            excel_img = ExcelImage(img_path)
            excel_img.width = crop_size
            excel_img.height = crop_size
            ws.add_image(excel_img, f"{thumbnail_column_letter}{row_idx}")
            ws.row_dimensions[row_idx].height = crop_size / 1.33
            ws.column_dimensions[thumbnail_column_letter].width = crop_size / 7
        except Exception:
            ws.cell(row=row_idx, column=8).value = "Image not found"

save_cache()
wb.save(output_excel)
print(f"Completed: {len(results)} new images processed out of {total_rows}")
